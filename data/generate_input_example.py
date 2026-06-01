"""
生成示例 input.xlsx
"""
import sys
sys.path.insert(0, r"E:\autoBackLine")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# ---- 样式定义 ----
header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=12)
header_fill = PatternFill("solid", fgColor="4472C4")
header_align = Alignment(horizontal="center", vertical="center")

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

data_align = Alignment(horizontal="center", vertical="center")

# ---- 表头 ----
headers = ["序号", "关键词", "域名"]
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# ---- 示例数据 ----
example_data = [
    [1, "SEO工具", "heartratetap.com"],
    [2, "网站分析", "example.com"],
    [3, "反向链接", "testsite.org"],
]

for row_idx, row_data in enumerate(example_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = data_align
        cell.border = thin_border

# ---- 列宽自适应 ----
col_widths = {"序号": 8, "关键词": 20, "域名": 25}
for col_idx, header in enumerate(headers, start=1):
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = col_widths.get(header, 15)

ws.row_dimensions[1].height = 25

# ---- 保存示例文件 ----
example_path = r"E:\autoBackLine\data\input_example.xlsx"
wb.save(example_path)
print(f"示例文件已生成: {example_path}")

# ---- 同时更新实际的 input.xlsx ----
input_path = r"E:\autoBackLine\data\input.xlsx"
wb.save(input_path)
print(f"已更新: {input_path}")
