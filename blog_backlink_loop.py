import argparse
import csv
import json
import os
import re
import shutil
import sys
import time
import ipaddress
from datetime import date
from pathlib import Path
from urllib.parse import urlparse

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

from core.backlinks_merger import load_all_blacklists
from extension_batch_upload import (
    CdpError,
    attach_target,
    existing_downloads,
    list_targets,
    new_target,
    request_json,
    unique_path,
    wait_until,
    write_event,
)


load_dotenv()

DEFAULT_BUTTON_TEXT = os.getenv("BLOG_ANALYSIS_BUTTON_TEXT", "导出外链")
DEFAULT_KEYWORD = os.getenv("BLOG_ANALYSIS_KEYWORD", "SEO工具")
URL_RE = re.compile(r"https?://[^\s\"'<>,，。；;]+", re.I)
DOMAIN_RE = re.compile(
    r"^(?:[a-z0-9](?:[a-z0-9-]{0,61}[a-z0-9])?\.)+[a-z]{2,}(?:/[^\s\"'<>,，。；;]*)?$",
    re.I,
)


def safe_filename(value):
    parsed = urlparse(ensure_url(value))
    host = parsed.netloc or parsed.path.split("/")[0] or "blog"
    path = parsed.path.strip("/").replace("/", "_")
    raw = f"{host}_{path}" if path else host
    return re.sub(r"[^A-Za-z0-9._-]+", "_", raw).strip("_")[:140] or "blog"


def ensure_url(value):
    text = str(value or "").strip()
    if not text:
        return ""
    if re.match(r"^https?://", text, re.I):
        return text
    return f"https://{text}"


def normalize_url(value):
    text = str(value or "").strip()
    if re.match(r"^[a-z][a-z0-9+.-]*://", text, re.I) and not re.match(r"^https?://", text, re.I):
        return ""
    url = ensure_url(text)
    parsed = urlparse(url)
    if parsed.scheme.lower() not in {"http", "https"} or not parsed.netloc:
        return ""
    scheme = parsed.scheme.lower() or "https"
    netloc = parsed.netloc.lower()
    path = parsed.path or "/"
    normalized = f"{scheme}://{netloc}{path}"
    if parsed.query:
        normalized += f"?{parsed.query}"
    return normalized


def normalize_domain(value):
    text = str(value or "").strip().lower()
    if not text:
        return ""
    if re.match(r"^[a-z][a-z0-9+.-]*://", text, re.I):
        parsed = urlparse(text)
        text = parsed.netloc
    text = text.split("/")[0].split("?")[0].split("#")[0].strip().lower()
    text = re.sub(r"^www\.", "", text)
    if not text or "." not in text:
        return ""
    return text


def newest_csv(directory, after_time=0):
    root = Path(directory)
    if not root.exists():
        return None
    files = [p for p in root.glob("*.csv") if p.stat().st_mtime >= after_time]
    if not files:
        files = list(root.glob("*.csv"))
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def read_rows(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader), reader.fieldnames or []


def select_blog_urls(batch_csv, result_filter):
    rows, _headers = read_rows(batch_csv)
    urls = []
    seen = set()
    for row in rows:
        row_type = str(row.get("类型", "")).strip()
        if row_type and row_type != "博客":
            continue

        result = str(row.get("运行结果", "")).strip()
        if result_filter == "success" and result != "√":
            continue
        if result_filter == "not-fail" and result == "×":
            continue

        raw_url = row.get("原URL") or row.get("URL") or row.get("url") or ""
        url = normalize_url(raw_url)
        if url and url not in seen:
            seen.add(url)
            urls.append(url)
    return urls


def set_download_dir(page, download_dir):
    try:
        page.send(
            "Browser.setDownloadBehavior",
            {"behavior": "allow", "downloadPath": str(Path(download_dir).resolve())},
        )
        return True
    except Exception as exc:
        write_event("Could not set Chrome download directory; will monitor fallback Downloads", error=str(exc))
        return False


def wait_ready(page, deadline):
    wait_until(deadline, lambda: page.eval("document.readyState !== 'loading'"), interval=0.5)


def find_target_by_url(port, url, timeout_seconds):
    for target in list_targets(port, timeout_seconds):
        if target.get("type") == "page" and target.get("url", "").startswith(url):
            return target
    return None


def open_page(port, url, timeout_seconds):
    target = find_target_by_url(port, url, timeout_seconds)
    if target:
        return attach_target(target, timeout_seconds)
    return attach_target(new_target(port, url, timeout_seconds), timeout_seconds)


def click_analysis_button(page, button_text, selectors, deadline):
    selector_list = json.dumps([s for s in selectors if s], ensure_ascii=False)
    button_text_json = json.dumps(button_text, ensure_ascii=False)
    expression = f"""(() => {{
      const selectors = {selector_list};
      const buttonText = {button_text_json};

      function isVisible(el) {{
        const style = getComputedStyle(el);
        const rect = el.getBoundingClientRect();
        return style.visibility !== 'hidden' && style.display !== 'none' && rect.width > 0 && rect.height > 0;
      }}

      function allRoots(root) {{
        const roots = [root];
        const nodes = root.querySelectorAll ? root.querySelectorAll('*') : [];
        for (const node of nodes) {{
          if (node.shadowRoot) roots.push(...allRoots(node.shadowRoot));
        }}
        return roots;
      }}

      function findInRoot(root) {{
        for (const selector of selectors) {{
          const el = root.querySelector && root.querySelector(selector);
          if (el && isVisible(el)) return {{ el, selector, method: 'selector' }};
        }}

        const candidates = root.querySelectorAll
          ? root.querySelectorAll('button, a, [role="button"], input[type="button"], input[type="submit"], div, span')
          : [];
        for (const el of candidates) {{
          const text = (el.innerText || el.textContent || el.value || '').trim();
          if (text.includes(buttonText) && isVisible(el)) {{
            return {{ el, selector: buttonText, method: 'text', text }};
          }}
        }}
        return null;
      }}

      for (const root of allRoots(document)) {{
        const found = findInRoot(root);
        if (found) {{
          found.el.scrollIntoView({{ block: 'center', inline: 'center' }});
          found.el.click();
          return {{
            ok: true,
            method: found.method,
            selector: found.selector,
            text: found.text || found.el.innerText || found.el.textContent || found.el.value || ''
          }};
        }}
      }}

      return {{
        ok: false,
        bodyText: document.body ? document.body.innerText.slice(0, 800) : ''
      }};
    }})()"""

    def click_once():
        result = page.eval(expression)
        if result and result.get("ok"):
            return result
        return None

    return wait_until(deadline, click_once, interval=1)


def stable_changed_csv(scan_dirs, before_maps, timeout_seconds):
    deadline = time.monotonic() + timeout_seconds
    last_candidate = None
    while time.monotonic() < deadline:
        for directory in scan_dirs:
            before = before_maps.get(str(directory), {})
            current = existing_downloads(str(directory))
            changed = []
            for path, signature in current.items():
                name = Path(path).name.lower()
                if name.endswith(".csv") and (path not in before or before[path] != signature):
                    changed.append(Path(path))
            if changed:
                changed.sort(key=lambda item: item.stat().st_mtime, reverse=True)
                candidate = changed[0]
                stat1 = candidate.stat()
                time.sleep(0.8)
                stat2 = candidate.stat()
                if stat1.st_size == stat2.st_size and stat1.st_mtime_ns == stat2.st_mtime_ns:
                    return candidate
                last_candidate = str(candidate)
        time.sleep(0.5)
    raise CdpError(f"Timed out waiting for analysis CSV download. Last candidate: {last_candidate}")


def analyze_blog_url(port, url, output_dir, button_text, selectors, timeout_seconds, download_timeout_seconds):
    fallback_dir = Path.home() / "Downloads"
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    scan_dirs = [output_path]
    if fallback_dir.exists() and fallback_dir.resolve() != output_path.resolve():
        scan_dirs.append(fallback_dir)
    before = {str(directory): existing_downloads(str(directory)) for directory in scan_dirs}

    page = open_page(port, url, timeout_seconds)
    try:
        set_download_dir(page, output_path)
        deadline = time.monotonic() + timeout_seconds
        wait_ready(page, deadline)
        click_result = click_analysis_button(page, button_text, selectors, deadline)
        if not click_result:
            raise CdpError(f"Could not find or click analysis button: {button_text}")
        write_event("Clicked blog analysis button", url=url, result=click_result)

        downloaded = stable_changed_csv(scan_dirs, before, download_timeout_seconds)
        target = unique_path(output_path / f"blog_analysis_{date.today().strftime('%Y-%m-%d')}_{safe_filename(url)}.csv")
        if downloaded.resolve() != target.resolve():
            shutil.move(str(downloaded), str(target))
        write_event("Blog analysis CSV saved", url=url, path=str(target.resolve()))
        return str(target.resolve())
    finally:
        page.close()


def extract_urls_from_csv(path):
    found = []
    with open(path, "r", encoding="utf-8-sig", newline="", errors="replace") as handle:
        sample = handle.read()
    for match in URL_RE.findall(sample):
        url = normalize_url(match)
        if url:
            found.append(url)

    with open(path, "r", encoding="utf-8-sig", newline="", errors="replace") as handle:
        reader = csv.reader(handle)
        for row in reader:
            for value in row:
                text = str(value or "").strip()
                if DOMAIN_RE.match(text):
                    url = normalize_url(text)
                    if url:
                        found.append(url)
    return found


def extract_outlink_records_from_csv(path):
    records = []
    with open(path, "r", encoding="utf-8-sig", newline="", errors="replace") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            return records
        for row in reader:
            raw_url = row.get("URL") or row.get("url") or row.get("Url") or ""
            url = normalize_url(raw_url)
            if not url:
                continue

            hostname = row.get("Hostname") or row.get("hostname") or row.get("Host") or ""
            domain = normalize_domain(hostname) or normalize_domain(url)
            if not domain:
                continue

            link_type = str(row.get("Type") or row.get("type") or "").strip()
            link_text = str(row.get("Link Text") or row.get("Text") or row.get("text") or "").strip()
            records.append({
                "url": url,
                "domain": domain,
                "type": link_type,
                "text": link_text,
                "source_csv": str(path),
            })
    return records


def is_blacklisted_outlink(record, url_blacklist, domain_blacklist):
    url = record["url"].lower()
    domain = record["domain"].lower()
    for keyword in url_blacklist:
        kw = str(keyword).strip().lower()
        if kw and kw in url:
            return True
    for keyword in domain_blacklist:
        kw = str(keyword).strip().lower()
        if kw and kw in domain:
            return True
    return False


def is_ip_domain(domain):
    value = str(domain or "").strip().strip("[]")
    try:
        ipaddress.ip_address(value)
        return True
    except ValueError:
        return False


def record_rank(record):
    link_type = str(record.get("type") or "").lower()
    return 1 if "dofollow" in link_type else 0


def merge_outlink_csvs(csv_paths, merged_csv_path):
    url_blacklist, domain_blacklist = load_all_blacklists()
    by_domain = {}
    raw_count = 0
    blacklisted_count = 0

    for path in csv_paths:
        for record in extract_outlink_records_from_csv(path):
            raw_count += 1
            if is_ip_domain(record["domain"]):
                blacklisted_count += 1
                continue
            if is_blacklisted_outlink(record, url_blacklist, domain_blacklist):
                blacklisted_count += 1
                continue
            domain = record["domain"]
            existing = by_domain.get(domain)
            if not existing or record_rank(record) > record_rank(existing):
                by_domain[domain] = record

    merged = sorted(by_domain.values(), key=lambda item: item["domain"])
    output = Path(merged_csv_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with open(output, "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["域名", "URL", "Type", "Link Text", "Source CSV"])
        for record in merged:
            writer.writerow([
                record["domain"],
                record["url"],
                record["type"],
                record["text"],
                record["source_csv"],
            ])

    return {
        "records": merged,
        "raw_count": raw_count,
        "blacklisted_count": blacklisted_count,
        "merged_csv": str(output.resolve()),
    }


def read_existing_keyword(input_xlsx, default_keyword):
    try:
        wb = load_workbook(input_xlsx, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        keyword_col = headers.index("关键词") + 1 if "关键词" in headers else 2
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row=row, column=keyword_col).value
            if value:
                return str(value)
    except Exception:
        pass
    return default_keyword


def overwrite_input_xlsx(input_xlsx, urls, keyword):
    path = Path(input_xlsx)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["序号", "关键词", "域名"])
    for index, url in enumerate(urls, 1):
        ws.append([index, keyword, url])
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 80
    wb.save(path)


def merge_analysis_csvs_to_input(csv_paths, input_xlsx, keyword, merged_csv_path=None):
    if merged_csv_path is None:
        merged_csv_path = Path(input_xlsx).parent / f"blog_outlinks_merged_{date.today().strftime('%Y-%m-%d')}.csv"

    result = merge_outlink_csvs(csv_paths, merged_csv_path)
    urls = [record["url"] for record in result["records"]]
    if not urls:
        raise RuntimeError("No URLs remained after blacklist filtering and domain de-duplication; input.xlsx was not overwritten.")

    overwrite_input_xlsx(input_xlsx, urls, keyword)
    result["urls"] = urls
    return result


def parse_args():
    parser = argparse.ArgumentParser(description="Analyze blog pages with the extension floating button and overwrite data/input.xlsx.")
    parser.add_argument("--batch-result", default="", help="Specific plugin batch result CSV.")
    parser.add_argument("--batch-result-dir", default="data/output", help="Directory containing plugin batch result CSV files.")
    parser.add_argument("--after-time", type=float, default=0, help="Only prefer batch result CSV files modified after this Unix timestamp.")
    parser.add_argument("--input-dir", default="data/input", help="Directory where blog analysis CSV files are saved.")
    parser.add_argument("--input-xlsx", default="data/input.xlsx", help="Workbook overwritten with newly exported URLs.")
    parser.add_argument("--merged-csv", default="", help="Merged outlink CSV path. Defaults to data/blog_outlinks_merged_YYYY-MM-DD.csv.")
    parser.add_argument("--port", type=int, default=9222, help="Chrome remote debugging port.")
    parser.add_argument("--button-text", default=DEFAULT_BUTTON_TEXT, help="Floating button text to click.")
    parser.add_argument("--button-selector", action="append", default=[], help="Additional CSS selector for the analysis button.")
    parser.add_argument("--result-filter", choices=["all", "success", "not-fail"], default="all", help="Which plugin result rows to analyze.")
    parser.add_argument("--max-pages", type=int, default=0, help="Optional cap for debugging.")
    parser.add_argument("--timeout-ms", type=int, default=45000, help="Page and click timeout.")
    parser.add_argument("--download-timeout-ms", type=int, default=60000, help="CSV download timeout after clicking analysis.")
    parser.add_argument("--keyword", default=DEFAULT_KEYWORD, help="Keyword value written to input.xlsx.")
    parser.add_argument("--dry-run", action="store_true", help="List blog URLs without opening pages.")
    return parser.parse_args()


def main():
    args = parse_args()
    timeout_seconds = max(args.timeout_ms / 1000, 1)
    download_timeout_seconds = max(args.download_timeout_ms / 1000, 1)

    batch_result = Path(args.batch_result) if args.batch_result else newest_csv(args.batch_result_dir, args.after_time)
    if not batch_result or not batch_result.exists():
        write_event("No plugin batch result CSV found", batch_result=args.batch_result, directory=args.batch_result_dir)
        return 2

    urls = select_blog_urls(batch_result, args.result_filter)
    if args.max_pages > 0:
        urls = urls[: args.max_pages]
    write_event("Selected blog URLs for analysis", batch_result=str(batch_result), count=len(urls))
    if not urls:
        return 0

    if args.dry_run:
        for url in urls:
            write_event("Dry-run blog URL", url=url)
        return 0

    try:
        request_json("GET", f"http://127.0.0.1:{args.port}/json/version", timeout_seconds)
        selectors = args.button_selector + [
            "[data-action='analyze-backlinks']",
            "[data-testid='analyze-backlinks']",
            "#analyzeBacklinks",
            ".analyze-backlinks",
        ]

        exported = []
        for index, url in enumerate(urls, 1):
            write_event("Analyzing blog page", index=index, total=len(urls), url=url)
            try:
                exported.append(
                    analyze_blog_url(
                        args.port,
                        url,
                        args.input_dir,
                        args.button_text,
                        selectors,
                        timeout_seconds,
                        download_timeout_seconds,
                    )
                )
            except Exception as exc:
                write_event("Blog analysis failed", url=url, error=str(exc))

        if not exported:
            write_event("No blog analysis CSV files were exported; input.xlsx was not overwritten")
            return 3

        keyword = args.keyword or read_existing_keyword(args.input_xlsx, DEFAULT_KEYWORD)
        merge_result = merge_analysis_csvs_to_input(exported, args.input_xlsx, keyword, args.merged_csv or None)
        write_event(
            "input.xlsx overwritten from blog analysis CSV files",
            input_xlsx=str(Path(args.input_xlsx).resolve()),
            csv_count=len(exported),
            raw_count=merge_result["raw_count"],
            blacklisted_count=merge_result["blacklisted_count"],
            domain_count=len(merge_result["urls"]),
            merged_csv=merge_result["merged_csv"],
        )
        return 0
    except Exception as exc:
        write_event("Blog backlink loop failed", error=str(exc))
        return 1


if __name__ == "__main__":
    sys.exit(main())
